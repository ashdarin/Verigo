from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path

import xlrd
from openpyxl import load_workbook


EMAIL_PATTERN = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")


def emails_in_values(values: list[object], limit: int) -> list[str]:
    emails: list[str] = []
    seen: set[str] = set()
    for value in values:
        for email in EMAIL_PATTERN.findall(str(value or "")):
            key = email.lower()
            if key not in seen:
                seen.add(key)
                emails.append(email)
                if len(emails) >= limit:
                    return emails
    return emails


def decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("文件编码无法识别，请使用 UTF-8 或 GB18030")


def extract_emails(filename: str, data: bytes, limit: int) -> list[str]:
    suffix = Path(filename).suffix.lower()
    values: list[object] = []

    if suffix == ".txt":
        values = [decode_text(data)]
    elif suffix == ".csv":
        reader = csv.reader(io.StringIO(decode_text(data)))
        values = [cell for row in reader for cell in row]
    elif suffix == ".json":
        payload = json.loads(decode_text(data))

        def flatten(value: object) -> None:
            if isinstance(value, dict):
                for item in value.values():
                    flatten(item)
            elif isinstance(value, list):
                for item in value:
                    flatten(item)
            else:
                values.append(value)

        flatten(payload)
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    values.extend(row)
        finally:
            workbook.close()
    elif suffix == ".xls":
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
        try:
            for sheet in workbook.sheets():
                for row_index in range(sheet.nrows):
                    values.extend(sheet.row_values(row_index))
        finally:
            workbook.release_resources()
    else:
        raise ValueError("支持 TXT、CSV、JSON、XLSX、XLSM、XLS 文件")

    return emails_in_values(values, limit)
