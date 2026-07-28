"""Parse user-owned company lists without relying on an external company database."""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from pathlib import Path

import xlrd
from openpyxl import load_workbook

from app.core.imports import decode_text
from app.core.prospecting import normalize_company_domain


@dataclass(frozen=True)
class ImportedCompany:
    name: str
    domain: str | None
    country: str | None
    industry: str | None
    source_row: int


HEADER_ALIASES = {
    "name": {"company", "company name", "company_name", "企业", "企业名称", "公司", "公司名称", "客户名称"},
    "domain": {"domain", "website", "site", "url", "官网", "网站", "网址", "域名"},
    "country": {"country", "国家", "所在国家", "国家地区"},
    "industry": {"industry", "sector", "行业", "所属行业", "赛道"},
}


def _header_map(row: list[object]) -> dict[str, int]:
    normalized = {str(value or "").strip().lower(): index for index, value in enumerate(row)}
    return {
        field: next((normalized[key] for key in aliases if key in normalized), -1)
        for field, aliases in HEADER_ALIASES.items()
    }


def _value(row: list[object], index: int) -> str:
    return str(row[index] or "").strip() if 0 <= index < len(row) else ""


def _parse_rows(rows: list[list[object]], limit: int) -> list[ImportedCompany]:
    if not rows:
        return []
    headers = _header_map(rows[0])
    if headers["name"] < 0 and headers["domain"] < 0:
        raise ValueError("Company list needs a company name or website/domain column")
    companies: list[ImportedCompany] = []
    seen: set[tuple[str, str]] = set()
    for source_row, row in enumerate(rows[1:], 2):
        name = _value(row, headers["name"])
        raw_domain = _value(row, headers["domain"])
        try:
            domain = normalize_company_domain(raw_domain) if raw_domain else None
        except ValueError:
            domain = None
        if not name and not domain:
            continue
        if not name:
            name = domain or ""
        key = (name.casefold(), domain or "")
        if key in seen:
            continue
        seen.add(key)
        companies.append(ImportedCompany(
            name=name[:200], domain=domain, country=_value(row, headers["country"])[:16] or None,
            industry=_value(row, headers["industry"])[:120] or None, source_row=source_row,
        ))
        if len(companies) >= limit:
            break
    return companies


def extract_companies(filename: str, data: bytes, limit: int = 5000) -> list[ImportedCompany]:
    suffix = Path(filename).suffix.lower()
    rows: list[list[object]] = []
    if suffix == ".csv":
        rows = [list(row) for row in csv.reader(io.StringIO(decode_text(data)))]
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows.extend([list(row) for row in sheet.iter_rows(values_only=True)])
        finally:
            workbook.close()
    elif suffix == ".xls":
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
        try:
            for sheet in workbook.sheets():
                rows.extend([list(sheet.row_values(index)) for index in range(sheet.nrows)])
        finally:
            workbook.release_resources()
    else:
        raise ValueError("Company lists support CSV, XLSX, XLSM, and XLS files")
    return _parse_rows(rows, limit)
